"""Utilities for exporting group data to Excel"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
import os
from datetime import datetime


def create_group_export(group, forms):
    """
    Create an Excel workbook with group data and images
    
    Args:
        group: Group object
        forms: List of Form objects
    
    Returns:
        BytesIO object containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Submissions"
    
    # Define styles
    header_fill = PatternFill(start_color="FF7A00", end_color="FF7A00", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    column_widths = {
        'A': 15,  # First Name
        'B': 15,  # Last Name
        'C': 15,  # Middle Name
        'D': 12,  # Eye Color
        'E': 12,  # Hair Color
        'F': 25,  # Address
        'G': 15,  # Date of Birth
        'H': 10,  # Height
        'I': 10,  # Weight
        'J': 8,   # State
        'K': 15,  # City
        'L': 12,  # Zip Code
        'M': 12,  # Gender
        'N': 12,  # Organ Donor
        'O': 20,  # Corrective Lenses
        'P': 20,  # Submitted At
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Create header row
    headers = [
        'First Name', 'Last Name', 'Middle Name', 'Eye Color', 'Hair Color',
        'Address', 'Date of Birth', 'Height (cm)', 'Weight (kg)', 'State',
        'City', 'Zip Code', 'Gender', 'Organ Donor', 'Corrective Lenses',
        'Submitted At'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Add data rows
    for row_num, form in enumerate(forms, 2):
        row_data = [
            form.first_name,
            form.last_name,
            form.middle_name or '',
            form.eye_color,
            form.hair_color,
            form.address or '',
            form.date_of_birth.strftime('%Y-%m-%d') if form.date_of_birth else '',
            form.height,
            form.weight,
            form.state or '',
            form.city or '',
            form.zip_code or '',
            form.gender,
            'Yes' if form.organ_donor else 'No',
            'Yes' if form.restrictions_corrective_lenses else 'No',
            form.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if form.submitted_at else '',
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            if col_num in [7, 8, 9]:  # Center align numeric columns
                cell.alignment = Alignment(horizontal='center')
    
    # Create a BytesIO object to store the workbook
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def get_image_filename(form):
    """
    Generate a standardized filename for a form's image
    Format: FirstName_LastName_id
    
    Args:
        form: Form object
    
    Returns:
        String filename without extension
    """
    return f"{form.first_name}_{form.last_name}_{form.id}"

